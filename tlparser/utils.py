import json
import openpyxl
import os
import pandas as pd
from datetime import datetime

from tlparser.config import Configuration
from tlparser.stats import Stats


class Utils:

    def __init__(self, config: Configuration):
        self.config = config

    def read_formulas_from_json(self):
        with open(self.config.file_data_in, "r") as file:
            data = json.load(file)
        parsed_formulas = []

        ids = [item["id"] for item in data]
        if len(ids) != len(set(ids)):
            raise ValueError("Duplicate IDs found, abort...")

        for entry in data:
            if entry["status"] in self.config.only_with_status:
                for logic in entry["logics"]:
                    s = Stats(logic["f_code"])
                    parsed_formulas.append(
                        {
                            "id": entry["id"],
                            "text": entry["text"],
                            "type": logic["type"],
                            "translation": logic["translation"],
                            "reasoning": logic["reasoning"],
                            "stats": s.get_stats(),
                        }
                    )
        return parsed_formulas

    def write_to_excel(self, data):
        flattened_data = [self.flatten_dict(item) for item in data]

        # Derive and append translation class
        df = pd.DataFrame(flattened_data)
        type_order = self.config.logic_order
        df["type"] = pd.Categorical(df["type"], categories=type_order, ordered=True)
        df_sorted = df.sort_values(by=["id", "type"])
        df["translationclass"] = df_sorted.groupby("id")["translation"].transform(
            lambda x: "".join([v[0] for v in x])
        )

        flattened_data = df.to_dict(orient="records")
        headers = {key for item in flattened_data for key in item.keys()}

        # Sort headers according to predefined order, with any extra headers at the end
        predefined_order = Utils.get_column_order()
        headers = [header for header in predefined_order if header in headers] + [
            header for header in headers if header not in predefined_order
        ]

        # Create a new workbook and select the active worksheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write the headers to the first row
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)

        # Write the data to the sheet
        for row, item in enumerate(flattened_data, start=2):
            for col, header in enumerate(headers, start=1):
                value = item.get(header)
                if (
                    value is None
                    or value == ""
                    or (isinstance(value, set) and len(value) == 0)
                ):
                    sheet.cell(row=row, column=col, value=None)
                elif isinstance(value, int) or isinstance(value, float):
                    sheet.cell(row=row, column=col, value=value)
                elif isinstance(value, set):
                    sheet.cell(row=row, column=col, value=" | ".join(value))
                else:
                    sheet.cell(row=row, column=col, value=str(value))

        # Save the workbook to a file
        os.makedirs(self.config.folder_data_out, exist_ok=True)
        prefix = self.extract_filename_without_suffix(self.config.file_data_in)
        out = os.path.join(
            self.config.folder_data_out, f"{prefix}_{self.get_unique_filename()}.xlsx"
        )
        workbook.save(out)
        return out

    @staticmethod
    def flatten_dict(d, parent_key="", sep="."):
        items = []
        for k, v in d.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(Utils.flatten_dict(v, new_key, sep=sep).items())
            else:
                items.append((new_key, v))
        return dict(items)

    @staticmethod
    def get_unique_filename():
        now = datetime.now()
        return now.strftime("%y%m%d%H%M%S")

    @staticmethod
    def extract_filename_without_suffix(file_path):
        base_name = os.path.basename(file_path)
        return os.path.splitext(base_name)[0]

    @staticmethod
    def get_latest_excel(folder):
        excel_files = [f for f in os.listdir(folder) if f.endswith(".xlsx")]
        if not excel_files:
            return ""

        # Sort files by modification time
        latest_file = max(
            excel_files, key=lambda f: os.path.getmtime(os.path.join(folder, f))
        )
        file = os.path.join(folder, latest_file)
        return file

    @staticmethod
    def get_column_order():
        return [
            "id",
            "text",
            "type",
            "reasoning",
            "translation",
            "translationclass",
            "stats.formula_raw",
            "stats.formula_parsable",
            "stats.formula_parsed",
            "stats.asth",
            "stats.ap",
            "stats.cops.eq",
            "stats.cops.geq",
            "stats.cops.gt",
            "stats.cops.leq",
            "stats.cops.lt",
            "stats.cops.neq",
            "stats.lops.and",
            "stats.lops.impl",
            "stats.lops.not",
            "stats.lops.or",
            "stats.tops.A",
            "stats.tops.E",
            "stats.tops.F",
            "stats.tops.G",
            "stats.tops.R",
            "stats.tops.U",
            "stats.tops.X",
            "stats.agg.aps",
            "stats.agg.cops",
            "stats.agg.lops",
            "stats.agg.tops",
            "stats.entropy.lops",
            "stats.entropy.tops",
            "stats.entropy.lops_tops",
        ]
